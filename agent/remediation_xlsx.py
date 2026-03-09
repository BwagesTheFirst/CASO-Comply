"""
CASO Comply -- Excel (.xlsx) Accessibility Remediation Engine

Analyzes and remediates Excel workbook accessibility issues using openpyxl
and lxml for OOXML manipulation.
"""
from __future__ import annotations

import asyncio
import logging
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.worksheet.table import Table
from lxml import etree

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 1. Analysis
# ---------------------------------------------------------------------------

def analyze_xlsx(file_path: str) -> dict:
    """Analyze Excel workbook for accessibility issues."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Check sheet names
    generic_names = {"Sheet", "Sheet1", "Sheet2", "Sheet3", "Sheet4", "Sheet5"}
    sheet_issues = [name for name in wb.sheetnames if name in generic_names]
    descriptive_sheets = len(wb.sheetnames) - len(sheet_issues)
    sheet_name_coverage = descriptive_sheets / max(len(wb.sheetnames), 1)

    # Check for merged cells
    total_merged = 0
    for ws in wb.worksheets:
        total_merged += len(ws.merged_cells.ranges)

    # Check tables for headers
    total_tables = 0
    tables_with_headers = 0
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            total_tables += 1
            if tbl.headerRowCount and tbl.headerRowCount > 0:
                tables_with_headers += 1

    table_header_coverage = tables_with_headers / max(total_tables, 1)

    # Check images for alt text
    total_images = 0
    images_with_alt = 0
    for ws in wb.worksheets:
        for img in ws._images:
            total_images += 1
            if hasattr(img, 'desc') and img.desc:
                images_with_alt += 1

    alt_text_coverage = images_with_alt / max(total_images, 1)

    # Check document properties
    has_title = bool(wb.properties.title)

    # Check language
    has_lang = _has_workbook_language(wb)

    # Check for color-only data (cells with fill but no text)
    color_only_cells = _count_color_only_cells(wb)

    # Compute score
    score = _compute_xlsx_score(
        table_header_coverage=table_header_coverage,
        total_merged=total_merged,
        sheet_name_coverage=sheet_name_coverage,
        alt_text_coverage=alt_text_coverage,
        has_lang=has_lang,
        has_title=has_title,
        total_images=total_images,
        total_tables=total_tables,
        color_only_cells=color_only_cells,
    )

    wb.close()

    return {
        "score": score,
        "sheets": {"total": len(wb.sheetnames), "generic_names": sheet_issues},
        "merged_cells": total_merged,
        "tables": {"total": total_tables, "with_headers": tables_with_headers},
        "images": {"total": total_images, "with_alt": images_with_alt},
        "has_title": has_title,
        "has_lang": has_lang,
        "color_only_cells": color_only_cells,
        "page_count": len(wb.sheetnames),
    }


def _has_workbook_language(wb) -> bool:
    """Check if workbook has a language property set."""
    return bool(wb.properties.language) if hasattr(wb.properties, 'language') else False


def _count_color_only_cells(wb) -> int:
    """Count cells that use fill color but have no text value (color-only info)."""
    count = 0
    NO_FILL = "00000000"
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 200)):
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb != NO_FILL and rgb != "00000000" and not cell.value:
                        count += 1
    return count


def _compute_xlsx_score(
    table_header_coverage: float,
    total_merged: int,
    sheet_name_coverage: float,
    alt_text_coverage: float,
    has_lang: bool,
    has_title: bool,
    total_images: int,
    total_tables: int,
    color_only_cells: int = 0,
) -> dict:
    """Compute 0-100 accessibility score for Excel workbook."""
    checks = {
        "table_headers": {
            "passed": table_header_coverage >= 0.9 if total_tables > 0 else True,
            "weight": 25,
            "description": "Tables have header rows" if total_tables > 0 else "No tables -- headers not needed",
        },
        "no_merged": {
            "passed": total_merged == 0,
            "weight": 15,
            "description": "No merged cells" if total_merged == 0 else f"{total_merged} merged cell ranges found",
        },
        "sheet_names": {
            "passed": sheet_name_coverage >= 0.9,
            "weight": 10,
            "description": "Sheet names are descriptive",
        },
        "alt_text": {
            "passed": alt_text_coverage >= 0.9 if total_images > 0 else True,
            "weight": 20,
            "description": "Images/charts have alt text" if total_images > 0 else "No images -- alt text not needed",
        },
        "language": {
            "passed": has_lang,
            "weight": 10,
            "description": "Workbook language specified",
        },
        "title": {
            "passed": has_title,
            "weight": 10,
            "description": "Workbook title set",
        },
        "no_color_only": {
            "passed": color_only_cells == 0,
            "weight": 10,
            "description": "No color-only information" if color_only_cells == 0 else f"{color_only_cells} cells use color without text",
        },
    }

    total = sum(c["weight"] for c in checks.values())
    earned = sum(c["weight"] for c in checks.values() if c["passed"])
    score = round(100 * earned / total) if total > 0 else 0

    return {"score": score, "checks": checks}


# ---------------------------------------------------------------------------
# 2. Remediation
# ---------------------------------------------------------------------------

def _unmerge_cells(wb):
    """Unmerge all merged cells and fill each cell with the original value."""
    for ws in wb.worksheets:
        merged_ranges = list(ws.merged_cells.ranges)
        for merge_range in merged_ranges:
            min_row = merge_range.min_row
            min_col = merge_range.min_col
            value = ws.cell(row=min_row, column=min_col).value

            ws.unmerge_cells(str(merge_range))

            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row=row, column=col).value = value


def _set_table_headers(wb):
    """Ensure all tables have header rows designated."""
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if not tbl.headerRowCount or tbl.headerRowCount == 0:
                tbl.headerRowCount = 1


# ---------------------------------------------------------------------------
# 3. Main Pipeline
# ---------------------------------------------------------------------------

async def remediate_xlsx_async(
    file_path: str,
    output_path: str | None = None,
    verify: bool = True,
) -> dict:
    """Full Excel workbook remediation pipeline."""
    file_path = str(file_path)
    if output_path is None:
        p = Path(file_path)
        output_path = str(p.parent / f"{p.stem}_remediated{p.suffix}")

    logger.info("Analyzing input XLSX: %s", file_path)
    before_analysis = analyze_xlsx(file_path)

    # Open workbook for remediation
    wb = openpyxl.load_workbook(file_path)

    # Apply heuristic fixes
    _unmerge_cells(wb)
    _set_table_headers(wb)

    # Set title if missing
    if not wb.properties.title:
        wb.properties.title = wb.sheetnames[0] if wb.sheetnames else "Untitled Workbook"

    # Set language if missing
    if not wb.properties.language:
        wb.properties.language = "en-US"

    # Gemini AI verification
    verification_info = None
    if verify:
        from gemini_verify import verify_and_correct_office

        logger.info("Running Gemini AI verification on XLSX")

        structure_data = {
            "sheets": wb.sheetnames,
            "tables_per_sheet": {},
            "merged_ranges_fixed": before_analysis["merged_cells"],
            "images_count": before_analysis["images"]["total"],
        }
        for ws in wb.worksheets:
            structure_data["tables_per_sheet"][ws.title] = [
                {"name": tbl.name, "ref": tbl.ref, "has_header": bool(tbl.headerRowCount)}
                for tbl in ws.tables.values()
            ]

        verification_result = await verify_and_correct_office(
            file_path=file_path,
            structure_data=structure_data,
            format="xlsx",
        )

        verification_info = {
            "issues_found": verification_result.get("issues_found", []),
            "verification_score": verification_result.get("verification_score", 0.0),
        }

        # Apply sheet name suggestions
        for suggestion in verification_result.get("sheet_names", []):
            original = suggestion.get("original")
            suggested = suggestion.get("suggested")
            if original and suggested and original in wb.sheetnames:
                wb[original].title = suggested

        # Apply alt texts
        for alt_entry in verification_result.get("alt_texts", []):
            sheet_name = alt_entry.get("sheet")
            obj_idx = alt_entry.get("object_index", 0)
            desc = alt_entry.get("description", "")
            if sheet_name and sheet_name in wb.sheetnames and desc:
                ws = wb[sheet_name]
                if obj_idx < len(ws._images):
                    img = ws._images[obj_idx]
                    if hasattr(img, 'desc'):
                        img.desc = desc

    # Save remediated workbook
    wb.save(output_path)
    wb.close()
    logger.info("Saved remediated XLSX: %s", output_path)

    # Analyze the output
    after_analysis = analyze_xlsx(output_path)

    result = {
        "before": before_analysis,
        "after": after_analysis,
        "output_path": output_path,
        "page_count": before_analysis.get("page_count", 1),
    }

    if verification_info:
        result["verification"] = verification_info

    return result
